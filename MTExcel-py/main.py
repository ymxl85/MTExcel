import pprint

import openpyxl
import re
import os
from openpyxl.formula import Tokenizer
from openpyxl.utils import column_index_from_string, get_column_letter
import MR_judge7 as MR
import time
import color_cells

'''

初步分析公式参考值的关系并初步区分
遍历每一块的公式并通过MR来得到每个公式所含有的MR关系
'''


# 获取excel文件目录
def file_name(path):
    for root, dirs, files in os.walk(path):
        return files


# 获取公式并解析
def is_formula(cell, cell_row, cell_col):
    def do_value(in_value):
        in_letters = re.findall('[A-Z]+', in_value)
        in_nums = re.findall('\d+', in_value)
        return in_letters, in_nums

    tok = Tokenizer(cell)
    arr = []
    ab_arr = []
    if '!' in cell:
        return arr, ab_arr
    for t in tok.items:
        if t.type is 'OPERAND' and t.subtype is 'RANGE':
            if '$' in t.value:
                ab_letters = re.findall('\$[A-Z]+\$\d+', t.value)
                if ab_letters:
                    for letter in ab_letters:
                        index = letter.find('$', letter.index('$') + 1)
                        t.value = t.value.replace(letter, '')
                        a = letter[:index]
                        a = a.replace('$', '')
                        a = int(column_index_from_string(a))
                        b = letter[index:]
                        b = int(b.replace('$', ''))
                        arr.append([b, a, 2])
                row_ab = re.findall('\$[A-Z]+\d+', t.value)
                if row_ab:
                    for row_letter in row_ab:
                        row = re.findall('[A-Z]+', row_letter)[0]
                        row = int(column_index_from_string(row))
                        column = int(re.findall('\d+', row_letter)[0])
                        ab_arr.append([column, row, 0])
                column_ab = re.findall('[A-Z]\$\d+', t.value)
                if column_ab:
                    for column_letter in column_ab:
                        row = re.findall('[A-Z]+', column_letter)[0]
                        row = int(column_index_from_string(row))
                        column = int(re.findall('\d+', column_letter)[0])
                        ab_arr.append([column, row, 1])
            elif ':' in t.value:
                letters, nums = do_value(t.value)
                if len(letters) != len(nums) or not letters:
                    break
                pre_letternum = int(column_index_from_string(letters[0]))
                post_letternum = int(column_index_from_string(letters[1]))
                pre_num = int(nums[0])
                post_num = int(nums[1])
                for i in range(pre_letternum, post_letternum + 1):
                    for j in range(pre_num, post_num + 1):
                        row_reference = int(j) - int(cell_row)
                        col_reference = int(i) - int(cell_col)
                        arr.append([row_reference, col_reference])
            else:
                letters, nums = do_value(t.value)
                if len(letters) != len(nums) or not letters:
                    break
                letter = letters[0]
                num = nums[0]
                letter_num = column_index_from_string(letter)
                row_reference = int(num) - int(cell_row)
                col_reference = int(letter_num) - int(cell_col)
                arr.append([row_reference, col_reference])
    return arr, ab_arr


# 对字典内分类
def slice_region(re_dict):
    k_list = []
    v_list = []
    for k, v in re_dict.items():
        v[0] = sorted(v[0])
        if v[0] in v_list:
            index = v_list.index(v[0])
            k_list[index].append(k)
        else:
            v_list.append(v[0])
            k_list.append([k])
    return k_list, v_list


def slice_2region(re_dict):
    k_list = []
    v_list = []
    for k, v in re_dict.items():
        if v in v_list:
            index = v_list.index(v)
            k_list[index].append(k)
        else:
            v_list.append(v)
            k_list.append([k])
    return k_list, v_list


def open_excel(excel_path):
    index = excel_path.rfind('.')
    str_list = list(excel_path)
    str_list.insert(index, '(6)')
    excel_path_fuben = ''.join(str_list)
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    final_sum = dict()
    number_list = [0 for i in range(7)]
    # 遍历每一个sheet
    for sheet in wb:
        # print(sheet.title)
        reference_dict = dict()
        rows = sheet.max_row
        cols = sheet.max_column
        # 遍历每一个单元格 得到每个单元格的参考值
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                cell = sheet.cell(row=row, column=col)
                # print(row,col,cell.value)
                if isinstance(cell.value, str):
                    # print('{}  {}'.format(row,col))
                    reference_arr, ab_reference = is_formula(cell.value, row, col)
                    if reference_arr:
                        reference_dict[(row, col)] = [reference_arr, ab_reference]
        # pprint.pprint(reference_dict)
        # 根据每个单元格的参考值来将他们区分开来
        region_list, reference_list = slice_region(reference_dict)
        # 根据每个单元格的绝对索引来二次区分
        region_abre = [[] for _ in range(len(region_list))]
        for i in range(len(region_list)):
            for cell in region_list[i]:
                region_abre[i].append(reference_dict[cell][1])
        # print(region_list)
        final_region = []
        final_reference = []
        for i in range(len(region_list)):
            # print(len(region_list))
            tag = 0
            for abre in region_abre[i]:
                if abre:
                    tag = 1
                    break
            if tag == 0:
                final_region.append(region_list[i])
                final_reference.append(reference_list[i])
                continue
            sub_region = []
            sub_re = []
            len_lists = []
            # print(region_list[i])
            # print(region_abre[i])
            for j in range(len(region_list[i])):
                if len(region_abre[i][j]) not in len_lists:
                    len_lists.append(len(region_abre[i][j]))
                    sub_re.append([])
                    index = len_lists.index(len(region_abre[i][j]))
                    sub_re[index].append(region_abre[i][j])
                    sub_region.append([region_list[i][j]])
                else:
                    index = len_lists.index(len(region_abre[i][j]))
                    sub_region[index].append(region_list[i][j])
                    sub_re[index].append(region_abre[i][j])
                # print(sub_region)
                # print(sub_re)
                # print(len_lists)
            other_reference = reference_list[i]
            for lenth in len_lists:
                if lenth == 0:
                    final_region.append(sub_region[len_lists.index(0)])
                    final_reference.append(other_reference)
                else:
                    new_list = []
                    for sub in sub_re[len_lists.index(lenth)]:
                        if sub not in new_list:
                            new_list.append(sub)
                    # print('-=-=-=-=-=-=-=')
                    # print(new_list)
                    # print(len(sub_re[len_lists.index(lenth)]))
                    # or len(new_list) / len(sub_re[len_lists.index(lenth)]) < 0.5
                    if len(new_list) == 1:
                        new_reference = reference_list[i].copy()
                        for n in range(len(sub_re[len_lists.index(lenth)][0])):
                            sub_re[len_lists.index(lenth)][0][n][2] = 2
                            new_reference.append(sub_re[len_lists.index(lenth)][0][n])
                        final_region.append(sub_region[len_lists.index(lenth)])
                        final_reference.append(new_reference)
                    else:
                        in_dict = dict()
                        keys = sub_region[len_lists.index(lenth)]
                        values = sub_re[len_lists.index(lenth)]
                        # print(keys)
                        # print(values)
                        for m in range(len(keys)):
                            key = keys[m]
                            value = values[m]
                            in_dict[key] = []
                            # print(value)
                            # print(key)
                            for n in range(len(value)):
                                row = value[n][0] - key[0]
                                colu = value[n][1] - key[1]
                                in_dict[key].append([row, colu])
                        in_regions, in_values = slice_2region(in_dict)
                        # print('------')
                        # print(in_dict)
                        # print(in_regions)
                        # print(in_values)
                        for in_region in in_regions:
                            final_region.append(in_region)
                        for in_value in in_values:
                            new_list = reference_list[i].copy()
                            for _ in in_value:
                                new_list.append(_)
                            final_reference.append(new_list)

        # print('region:')
        # for _ in final_region:
        #     print(_)
        # # print('=============')
        # print('reference:')
        # for _ in final_reference:
        #     print(_)
        # print(final_region)
        # print(final_reference)
        # print(len(final_region))
        # print(len(final_reference))
        finally_region = []
        finally_reference = []
        for i in range(len(final_region)):
            big_region = final_region[i]
            sub_reference = final_reference[i]
            no_absolute_reference = [sub_reference[i] for i in range(len(sub_reference)) if len(sub_reference[i])!=3]
            if len(no_absolute_reference) == 1 or len(big_region) == 1:
                finally_region.append(big_region)
                finally_reference.append(sub_reference)
                continue
            row_list = []
            column_list = []
            # 获取reference的行列数
            for sub_sub_reference in sub_reference:
                if len(sub_sub_reference) != 3:
                    row_list.append(sub_sub_reference[0])
                    column_list.append(sub_sub_reference[1])
            row_list = list(set(row_list))
            column_list = list(set(column_list))
            if len(row_list) > len(column_list):
                tag = 0
            else:
                tag = 1
            # 获取对应行或者列的数量
            set_list = []
            for cell in big_region:
                set_list.append(cell[tag])
            set_list = list(set(set_list))
            if len(set_list) == 1:
                finally_region.append(big_region)
                finally_reference.append(sub_reference)
                continue
            small_region = []
            for _ in range(len(set_list)):
                small_region.append([])
            for cell in big_region:
                index = set_list.index(cell[tag])
                small_region[index].append(cell)
            for smaller_region in small_region:
                finally_region.append(smaller_region)
                finally_reference.append(sub_reference)
        # print(finally_region)
        # print(finally_reference)
        # print(len(finally_region))
        # print(len(finally_reference))
        # for _ in region_abre:
        #     print(_)
        if final_region:
            final_region, final_mrs, sub_reference, suspicion, swap_sus, result_number_list = MR.find_mr(excel_path,
                                                                                                         excel_path_fuben,
                                                                                                         sheet.title,
                                                                                                         finally_region,
                                                                                                         finally_reference)
            final_sum[sheet.title] = []
            final_sum[sheet.title].append(final_region)
            final_sum[sheet.title].append(final_mrs)
            final_sum[sheet.title].append(sub_reference)
            final_sum[sheet.title].append(suspicion)
            final_sum[sheet.title].append(swap_sus)
            number_list = [number_list[x] + result_number_list[x] for x in range(7)]
        # pprint.pprint(final_sum)
        color_cells.color_cell(final_sum, excel_path,
                               excel_path_fuben)
    return number_list
    # with open('result3.txt', 'a+') as f:
    #     for k, v in final_sum.items():
    #         if v:
    #             if v[3]:
    #                 f.write(excel_path + ':\n')
    #                 f.write(k + ':')
    #                 for cell, sus in v[3].items():
    #                     if sus <= 0.01:
    #                         continue
    #                     row = cell[0]
    #                     column = get_column_letter(cell[1])
    #                     f.write(column + str(row) + ',')
    #                 f.write('\n')
    #     f.write('\n\n')


from win32com.client import DispatchEx
import pythoncom

pythoncom.CoInitialize()
xlApp = DispatchEx("Excel.Application")
xlApp.Visible = False
xlApp.DisplayAlerts = False
start = time.time()
path = r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx'
excel_list = file_name(path)
wrong_list = []
# wb = openpyxl.load_workbook('result.xlsx')
# ws = wb.active
# nrows = ws.max_row
# for i in range(2,nrows+1):
#     if ws.cell(row=i,column=6).value is not None:
#         try:
#             open_excel(ws.cell(row=i,column=1).value)
#         except:
#             wrong_list.append(i)
f = open('result_number1.txt', 'a+')
f_excel = open('result_excel.txt','a+')

EUSES_path = r'E:\pydate\EXCEL\EUSES_modified\EUSES\spreadsheets'
print(os.listdir(EUSES_path))
mutant_list = [0 for i in range(7)]
for dir_name in os.listdir(EUSES_path)[1:-1]:
    excel_dir = os.path.join(EUSES_path, dir_name + '\\SEEDED\\xlsx')
    excel_mut_list = file_name(excel_dir)
    for excel_file in excel_mut_list:
        print(excel_file)
        if not excel_file.endswith(').xlsx'):
            excel_path = os.path.join(excel_dir, excel_file)
            try:
                result_numberlist = open_excel(excel_path)
                mutant_list = [mutant_list[x] + result_numberlist[x] for x in range(7)]
                f_excel.write(excel_path+':'+str(result_numberlist)+'\n')
                print(mutant_list)
            except:
                wrong_list.append(excel_path)

f.write(str(mutant_list))
# reallife_list = [0 for i in range(7)]
# #
# for excel_file in excel_list:
#
#     if excel_file.endswith('xlsx') and not excel_file.endswith(').xlsx'):
#
#         excel_path = "{}{}{}".format(path, '\\', excel_file)
#         # print(excel_path)
#         try:
#             result_numberlist = open_excel(excel_path)
#             reallife_list = [reallife_list[x] + result_numberlist[x] for x in range(7)]
#             f_excel.write(excel_path+':'+str(result_numberlist))
#             # print(result_numberlist)
#             # print(reallife_list)
#         except:
#             print('{} is wrong!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'.format(excel_path))
#             wrong_list.append(excel_path)
# f.write(str(reallife_list))
f.close()
f_excel.close()
# open_excel(r'E:\pydate\EXCEL\custodes/Default original spreadsheets/xlsx/01-38-PK_tables-figures.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\3763250_Q304_factsheet.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\2003-4%20budget.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\act3_lab23_posey.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\Ag%20Statistics,%20NUE_2003.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\am_skandia_fin_supple#A80EE.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\Annexure%20(Audited%2#A7E05.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\DDAA_HW.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\eg_spreadsheets.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\document_de_reference#A828A.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\fin_accounts.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\grades_Spring04_Geol%#A8A32.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\Lalit_TimeReport_Fall02.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\Unaudited%20Dec%2003.xlsx')
# open_excel(r'E:\pydate\EXCEL\custodes\Default original spreadsheets\xlsx\Sponsoredprograms.xlsx')
# open_excel(r'E:\pydate\EXCEL\test15.xlsx')
# number=open_excel(r'E:\pydate\EXCEL\EUSES_modified\EUSES\spreadsheets\database\SEEDED\xlsx\DB_Admin_1FAULTS_FAULTVERSION2.xlsx')
# print(number)
end = time.time()
print(end - start)
print(wrong_list)
xlApp.Quit()
del (xlApp)
