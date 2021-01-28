import openpyxl
from openpyxl.styles import PatternFill
from win32com.client import Dispatch
from openpyxl.comments import Comment
import pythoncom
from openpyxl.utils import get_column_letter


def just_open(filename):
    pythoncom.CoInitialize()
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlApp.DisplayAlerts = False
    try:
        xlBook = xlApp.Workbooks.Open(filename)
        xlBook.Save()
        xlBook.Close()
    except:
        pass
    finally:
        xlApp.Quit()


def color_cell(final_sum, filename, filename_fuben):
    wb = openpyxl.load_workbook(filename, data_only=False)
    color_list = ['1874CD',
                  'FF3030',
                  'EEEE00',
                  'EE7621',
                  'BCEE68',
                  '9400D3',
                  '8B4513',
                  '00FFFF',
                  '556B2F',
                  'EE00EE',
                  'B22222',
                  'EE9A00',
                  'FA8072',
                  'DB7093',
                  'B9D3EE',
                  '8B864E',
                  '7B68EE',
                  'EEB422',
                  '00CD00',
                  '4D4D4D',
                  '8B7500',
                  '7F7F7F',
                  'FF6600',
                  '800000',
                  '008080',
                  'FFFFCC',
                  'CCFFCC',
                  'CCCCFF',
                  'FFCC99',
                  'CC99FF',
                  '000080',
                  '666699',
                  '969696',
                  '339966',
                  '993300',
                  '333300',
                  '99CCFF',
                  'C0C0C0',
                  '808080',
                  '9999FF',
                  '993366',
                  '008000',
                  '000080',
                  '800080',
                  '808080',
                  'FF00FF'

                  ]
    sus_color = 'FF00FF'
    fill_sus = PatternFill("lightGrid", fgColor=sus_color)
    for k, v in final_sum.items():

        ws = wb[k]
        rows = ws.max_row
        suspicion = v[3]
        swap_sus = v[4]
        # print(v[3])
        suspicion_text = 'suspicious:'
        routate_false = []
        for key in sorted(suspicion, key=suspicion.__getitem__, reverse=True):
            value = suspicion[key]

            if value > 0.01:
                sus_row = key[0]
                sus_column = get_column_letter(key[1])
                suspicion_text = suspicion_text + '{}{},  '.format(sus_column, sus_row)
                # ws.cell(key[0],key[1]).fill = fill_sus
        ws.cell(row=rows + 1, column=1).value = suspicion_text
        for i in range(len(v[0])):
            # print(i)
            fill = PatternFill("solid", fgColor=color_list[i % len(color_list)])
            mrs = v[1][i]
            reference = v[2][i]
            # mr_comment= ''
            # for mr in mrs:
            #     mr_comment.join()
            # comment = Comment('''refenerence :{}\n\nmrs :{}
            # '''.format(reference, mrs), 'mr')
            for cell in v[0][i]:
                comment = 'reference:'
                for j, subreference in enumerate(reference):
                    # print('=========')
                    # print(subreference)
                    if len(subreference) == 2:
                        subreference_row = subreference[0] + cell[0]
                        subreference_column = get_column_letter(subreference[1] + cell[1])
                    else:
                        subreference_row = subreference[0]
                        subreference_column = get_column_letter(subreference[1])
                    if j == len(reference) - 1:
                        comment = comment + '{}{}'.format(subreference_column, subreference_row)
                    else:
                        comment = comment + '{}{}'.format(subreference_column, subreference_row) + ','
                comment = comment + '\n\n' + 'mrs:'
                # print(mrs[0])
                if mrs:
                    if mrs[0] is True or mrs[0] is False:
                        comment = comment + '\nforward:{}\n'.format(mrs[0])
                        # print(isinstance(mrs[1],dict))
                        for key, values in mrs[1].items():
                            comment = comment + '{}:'.format(key)
                            for value in values:
                                comment = comment + '('
                                for sub_value in value:
                                    if len(sub_value[0]) == 3:
                                        value_reference_row = sub_value[0][0]
                                        value_reference_column = get_column_letter(sub_value[0][1])
                                    else:
                                        value_reference_row = sub_value[0][0] + cell[0]
                                        value_reference_column = get_column_letter(sub_value[0][1] + cell[1])
                                    if len(sub_value[-1]) == 3:
                                        value_reference_last_row = sub_value[-1][0]
                                        value_reference_last_column = get_column_letter(sub_value[-1][1])
                                    else:
                                        value_reference_last_row = sub_value[-1][0] + cell[0]
                                        value_reference_last_column = get_column_letter(sub_value[-1][1] + cell[1])
                                    comment = comment + '{}{}:{}{},'.format(value_reference_column, value_reference_row,
                                                                            value_reference_last_column,
                                                                            value_reference_last_row)
                                comment = comment + ')\n'
                        comment = comment + '\n' + 'add:'
                        for key, value in mrs[2].items():
                            comment = comment + key + ':' + str(value) + '\n'
                    else:
                        for mr in mrs:
                            # print(mr[1])
                            if mr[1] is '+':
                                if len(mr[0]) == 3:
                                    reference_row = mr[0][0]
                                    referrer_column = get_column_letter(mr[0][1])
                                else:
                                    # print(cell[0])
                                    # print(mr[0][0])
                                    reference_row = mr[0][0] + cell[0]
                                    referrer_column = get_column_letter(mr[0][1] + cell[1])
                                comment = comment + '({}{},+,{})  '.format(referrer_column, reference_row, mr[2])
                                # comment = comment + '('
                                # for references in mr[0]:
                                #     # print(references)
                                #     if len(references) == 3:
                                #         reference_row = references[0]
                                #         reference_column = get_column_letter(references[1])
                                #     else:
                                #         reference_row = references[0] + cell[0]
                                #         reference_column = get_column_letter(references[1] + cell[1])
                                #     comment = comment + '{}{},'.format(reference_column, reference_row)
                                # comment = comment + '+,{})  '.format(mr[2])
                            if mr[1] is '*':
                                comment = comment + '('
                                for references in mr[0]:
                                    # print(references)
                                    if len(references) == 3:
                                        reference_row = references[0]
                                        reference_column = get_column_letter(references[1])
                                    else:
                                        reference_row = references[0] + cell[0]
                                        reference_column = get_column_letter(references[1] + cell[1])
                                    comment = comment + '{}{},'.format(reference_column, reference_row)
                                comment = comment + '*,{})  '.format(mr[2])

                # print(comment)
                comment = comment + '\n\n' + 'Rotate:'
                if (cell[0], cell[1]) not in swap_sus:
                    comment = comment + 'True'
                else:
                    comment = comment + 'False'
                comment = Comment(comment, 'mr')
                ws.cell(row=cell[0], column=cell[1]).fill = fill
                ws.cell(row=cell[0], column=cell[1]).comment = comment
        for key, value in suspicion.items():
            if value > 0.01:
                ws.cell(key[0], key[1]).fill = fill_sus

    wb.save(filename_fuben)
    just_open(filename_fuben)
