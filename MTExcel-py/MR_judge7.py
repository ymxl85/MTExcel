import copy
import re

from win32com.client import DispatchEx
import openpyxl
import itertools
import pythoncom
import xlrd
from openpyxl.utils import get_column_letter
import random
import os

pythoncom.CoInitialize()
xlApp = DispatchEx("Excel.Application")
xlApp.Visible = False
xlApp.DisplayAlerts = False


def check_swap(ws_true, filename, sub_region, sheetname, in_wb, sheetname_num, forward):
    ws2 = in_wb.sheet_by_name(str(sheetname_num))
    different_index = []
    count = 0
    if forward == 1:
        for i in range(len(sub_region)):
            if i == 0:
                cell1_row = sub_region[i][0]
                cell1_column = sub_region[i][1]
                cell2_row = sub_region[-1][0]
                cell2_column = sub_region[-1][1]
            else:
                cell1_row = sub_region[i][0]
                cell1_column = sub_region[i][1]
                cell2_row = sub_region[i - 1][0]
                cell2_column = sub_region[i - 1][1]
            if ws2.cell(cell1_row - 1, cell1_column - 1).ctype == 2:
                number1 = ws2.cell(cell1_row - 1, cell1_column - 1).value
            else:
                number1 = 0
            if ws_true.cell(cell2_row - 1, cell2_column - 1).ctype == 2:
                number2 = ws_true.cell(cell2_row - 1, cell2_column - 1).value
            else:
                number2 = 0
            if abs(number1 - number2) > 0.0001:
                different_index.append((cell1_row, cell1_column))
                count += 1

    if forward == 0:
        for i in range(len(sub_region)):
            if i == len(sub_region) - 1:
                cell1_row = sub_region[i][0]
                cell1_column = sub_region[i][1]
                cell2_row = sub_region[0][0]
                cell2_column = sub_region[0][1]
            else:
                cell1_row = sub_region[i][0]
                cell1_column = sub_region[i][1]
                cell2_row = sub_region[i + 1][0]
                cell2_column = sub_region[i + 1][1]
            if ws2.cell(cell1_row - 1, cell1_column - 1).ctype == 2:
                number1 = ws2.cell(cell1_row - 1, cell1_column - 1).value
            else:
                number1 = 0
            if ws_true.cell(cell2_row - 1, cell2_column - 1).ctype == 2:
                number2 = ws_true.cell(cell2_row - 1, cell2_column - 1).value
            else:
                number2 = 0
            if abs(number1 - number2) > 0.0001:
                different_index.append((cell1_row, cell1_column))
                count += 1
    if count > 0.8 * len(sub_region):
        # print(1111)
        return []
    # print('=-=-===========-=--=-=-=-=', different_index)
    return different_index


def swap_excel(sub_region, sub_reference, ws_true, sheetname, wb, filename_fuben, ws_false, in_wb, in_sheetname,
               forward, swap_cell_list):
    if str(in_sheetname) not in in_wb.sheetnames:
        in_ws = in_wb.create_sheet(str(in_sheetname))
    else:
        in_ws = in_wb[str(in_sheetname)]

    if forward == 1:
        for i in range(len(sub_region)):
            if i == 0:
                for sub_sub_reference in sub_reference:
                    if len(sub_sub_reference) != 3:
                        cell1_row = sub_region[i][0] + sub_sub_reference[0]
                        cell1_column = sub_region[i][1] + sub_sub_reference[1]
                        cell2_row = sub_region[-1][0] + sub_sub_reference[0]
                        cell2_column = sub_region[-1][1] + sub_sub_reference[1]
                        in_ws.cell(row=cell1_row, column=cell1_column).value = ws_true.cell(cell2_row - 1,
                                                                                            cell2_column - 1).value
                    else:
                        cell_row = sub_sub_reference[0]
                        cell_column = sub_sub_reference[1]
                        if (cell_row, cell_column) in swap_cell_list:
                            return 0
                        else:
                            in_ws.cell(row=cell_row, column=cell_column).value = ws_true.cell(cell_row - 1,
                                                                                              cell_column - 1).value
            else:
                for sub_sub_reference in sub_reference:
                    if len(sub_sub_reference) != 3:
                        cell1_row = sub_region[i][0] + sub_sub_reference[0]
                        cell1_column = sub_region[i][1] + sub_sub_reference[1]
                        cell2_row = sub_region[i - 1][0] + sub_sub_reference[0]
                        cell2_column = sub_region[i - 1][1] + sub_sub_reference[1]
                        in_ws.cell(row=cell1_row, column=cell1_column).value = ws_true.cell(cell2_row - 1,
                                                                                            cell2_column - 1).value
                    else:
                        cell_row = sub_sub_reference[0]
                        cell_column = sub_sub_reference[1]
                        if (cell_row, cell_column) in swap_cell_list:
                            return 0
                        else:
                            in_ws.cell(row=cell_row, column=cell_column).value = ws_true.cell(cell_row - 1,
                                                                                              cell_column - 1).value
    if forward == 0:
        for i in range(len(sub_region)):
            if i == len(sub_region) - 1:
                for sub_sub_reference in sub_reference:
                    if len(sub_sub_reference) != 3:

                        cell1_row = sub_region[i][0] + sub_sub_reference[0]
                        cell1_column = sub_region[i][1] + sub_sub_reference[1]
                        cell2_row = sub_region[0][0] + sub_sub_reference[0]
                        cell2_column = sub_region[0][1] + sub_sub_reference[1]
                        in_ws.cell(row=cell1_row, column=cell1_column).value = ws_true.cell(cell2_row - 1,
                                                                                            cell2_column - 1).value
                    else:
                        cell_row = sub_sub_reference[0]
                        cell_column = sub_sub_reference[1]
                        if (cell_row, cell_column) in swap_cell_list:
                            return 0
                        else:
                            in_ws.cell(row=cell_row, column=cell_column).value = ws_true.cell(cell_row - 1,
                                                                                              cell_column - 1).value
            else:
                for sub_sub_reference in sub_reference:
                    if len(sub_sub_reference) != 3:

                        cell1_row = sub_region[i][0] + sub_sub_reference[0]
                        cell1_column = sub_region[i][1] + sub_sub_reference[1]
                        cell2_row = sub_region[i + 1][0] + sub_sub_reference[0]
                        cell2_column = sub_region[i + 1][1] + sub_sub_reference[1]
                        in_ws.cell(row=cell1_row, column=cell1_column).value = ws_true.cell(cell2_row - 1,
                                                                                            cell2_column - 1).value
                    else:
                        cell_row = sub_sub_reference[0]
                        cell_column = sub_sub_reference[1]
                        if (cell_row, cell_column) in swap_cell_list:
                            return 0
                        else:
                            in_ws.cell(row=cell_row, column=cell_column).value = ws_true.cell(cell_row - 1,
                                                                                              cell_column - 1).value
    for m in range(len(sub_region)):
        in_ws.cell(row=sub_region[m][0], column=sub_region[m][1]).value = ws_false.cell(row=sub_region[m][0],
                                                                                        column=sub_region[m][
                                                                                            1]).value
    return 1


def final_check(final_region_list, suspicion, ws):
    def get_side_gap_list(region_list):
        rows = []
        columns = []
        for cell in region_list:
            row = cell[0]
            column = cell[1]
            rows.append(row)
            columns.append(column)
        rows = list(set(rows))
        columns = list(set(columns))
        rows_length = len(rows)
        columns_length = len(columns)
        if rows_length > columns_length:
            tag = 0
        else:
            tag = 1
        region_dict = dict()
        sum_row_column = [columns, rows]
        for short in sum_row_column[tag]:
            middle_list = []
            for cell in region_list:
                if cell[abs(tag - 1)] == short:
                    middle_list.append(cell[tag])
            region_dict[short] = middle_list
        side_answer = []
        gap_answer = []
        len_gap_answer_frined = []
        for key, value in region_dict.items():
            result = []
            min_value = min(value)
            max_value = max(value)
            check = [i for i in range(min_value, max_value + 1)]
            answers = list(set(check) - set(value))
            if tag == 1:
                min_answer = (key, min_value - 1)
                if max_value < ws.ncols:
                    max_answer = (key, max_value + 1)
                else:
                    max_answer = -1
            else:
                min_answer = (min_value - 1, key)
                if max_value < ws.nrows:
                    max_answer = (max_value + 1, key)
                else:
                    max_answer = -1
            if min_value - 1 > 0:
                if ws.cell_type(min_answer[0] - 1, min_answer[1] - 1) == 2:
                    side_answer.append(min_answer)
            if max_answer != -1:
                if ws.cell_type(max_answer[0] - 1, max_answer[1] - 1) == 2:
                    side_answer.append(max_answer)
            for answer in answers:
                if tag == 1:
                    if ws.cell_type(key - 1, answer - 1) == 2:
                        result.append((key, answer))
                else:
                    if ws.cell_type(answer - 1, key - 1) == 2:
                        result.append((answer, key))
            len_gap_answer_frined.append(len(value))
            gap_answer.append(result)

        return side_answer, gap_answer, len_gap_answer_frined, tag

    def find_referenced_list(final_region_list, cell):
        cells_list_dict = []
        left_cells = cell
        cells_len = []
        tag_list = []
        for cells in final_region_list:
            same_cells = list(set(left_cells) & set(cells))
            if same_cells:
                _, _, _, tag = get_side_gap_list(cells)
                tag_list.append(tag)
                left_cells = list(set(left_cells) - set(same_cells))
                cells_list_dict.append(same_cells)
                cells_len.append(len(cells))
            if not left_cells:
                return cells_list_dict, cells_len, tag_list
        cells_list_dict.append(left_cells)
        cells_len.append(0)
        return cells_list_dict, cells_len, tag_list

    for region_list in final_region_list:
        if len(region_list) == 1:
            continue
        side_answer, gap_answer, len_gap_answer_friend, tag = get_side_gap_list(region_list)
        for side_cell in side_answer:
            if side_cell in suspicion.keys():
                continue
            _ = []
            _.append(side_cell)
            result_len, len_list, tag_list = find_referenced_list(final_region_list, _)
            if len_list[0] == 1:
                suspicion[side_cell] = 0.3
            if len_list[0] < len(region_list) * 0.25 and 0 < len_list[0] < 5 and tag_list[0] == tag:
                suspicion[side_cell] = 0.3 / len_list[0]
            if len_list[0] == 0:
                suspicion[side_cell] = 0.01
        for i, gap_cells in enumerate(gap_answer):
            if len(gap_cells) / len_gap_answer_friend[i] > 0.5:
                # print('=============', gap_cells)
                continue
            final_gap_cells = []
            for gap_cell in gap_cells:
                if gap_cell not in suspicion.keys():
                    final_gap_cells.append(gap_cell)
            suspicion_socre = 1 - (len(final_gap_cells) / len_gap_answer_friend[i])
            if final_gap_cells:
                result_len_dict, gap_len_list, tag_list = find_referenced_list(final_region_list, final_gap_cells)
                for m in range(len(result_len_dict)):
                    suspicion_socre_gap = suspicion_socre
                    if gap_len_list[m] > len(result_len_dict[m]):
                        if tag_list[m] == tag:
                            suspicion_socre_gap = suspicion_socre - (0.2 * (gap_len_list[m] - len(result_len_dict[m])))
                        else:
                            continue
                    if gap_len_list[m] == 0:
                        suspicion_socre_gap = suspicion_socre - (0.2 * len(result_len_dict[m]))
                        for gap_contast in result_len_dict[m]:
                            suspicion[gap_contast] = suspicion_socre_gap
                        continue
                    for gap_cell in result_len_dict[m]:
                        suspicion[gap_cell] = suspicion_socre_gap


def find_suspicion(regions, mrs):
    suspicion = []
    suspicion_num = []
    len_dict = dict()
    for i, mr in enumerate(mrs):
        if len(mr) in len_dict.keys():
            len_dict[len(mr)].append(i)
        else:
            len_dict[len(mr)] = []
            len_dict[len(mr)].append(i)
    for k, values in len_dict.items():
        length = 0
        for value in values:
            length = len(regions[value]) + length
        len_dict[k].append(length)
    for values in len_dict.values():
        for value in values[:-1]:
            if len(regions[value]) / values[-1] < 0.34:
                suspicion.append(regions[value])

                suspicion_num.append((1 - (len(regions[value]) / values[-1])) * 1000)
    return suspicion, suspicion_num


def mr_suspicion(region, list_a):
    def get_jaccard(a, b):
        c = [val for val in a if val in b]
        # print(c)
        # bingji
        aa = copy.deepcopy(a)
        for val in b:
            if val not in c:
                aa.append(val)
        jaccard = float(len(c) / len(aa))
        # print(jaccard)
        return jaccard

    sus_dict = dict()
    for i, a in enumerate(list_a):
        if not a or a[0] is True or a[0] is False:
            continue
        for j, b in enumerate(list_a[i + 1:]):
            # print(i)
            # print(b)
            if not b or b[0] is True or b[0] is False:
                continue
            # print(a)
            # print(b)
            jac = get_jaccard(a, b)
            # print('{}-{}\n{}'.format(j + i + 1, i, jac))
            if jac != float(0):
                if len(region[i]) < len(region[j + i + 1]):
                    index = i
                    k = pow(len(region[i]), 2) / len(region[j + i + 1])
                else:
                    index = i + j + 1
                    k = pow(len(region[j + i + 1]), 2) / len(region[i])
                # print(jac)
                # print(region[index])
                for sub_sus_region in region[index]:
                    if jac >= 0.5:
                        if sub_sus_region not in sus_dict.keys():
                            sus_dict[sub_sus_region] = jac / k
                        else:
                            if sus_dict[sub_sus_region] < jac / k:
                                sus_dict[sub_sus_region] = jac / k
                    # else:
                    #     if sub_sus_region not in sus_dict.keys():
                    #         sus_dict[sub_sus_region] = jac
                    #     else:
                    #         if sus_dict[sub_sus_region] < jac:
                    #             sus_dict[sub_sus_region] = jac
    return sus_dict


def check_multimr(ws1, filename2, regions, reference, sheetname, buffer, average_buffer, wb2, in_sheetname):
    ws2 = wb2.sheet_by_name(str(in_sheetname))
    # wb2 = openpyxl.load_workbook(filename2, data_only=True)
    # ws2 = wb2[sheetname]
    check_out = []
    for single in regions:
        cell_row = single[0]
        cell_col = single[1]
        if (cell_row, cell_col) in average_buffer.keys():
            re_cell = average_buffer[(cell_row, cell_col)]
        else:
            try:
                re_cell = float(ws1.cell_value(cell_row - 1, cell_col - 1)) if ws1.cell_value(cell_row - 1,
                                                                                              cell_col - 1) is not None else 1
            except:
                re_cell = 0
        if (cell_row, cell_col) in buffer.keys():
            follow_cell = float(buffer[cell_row, cell_col])
        else:
            try:
                follow_cell = float(ws2.cell_value(cell_row - 1, cell_col - 1)) if ws2.cell_value(cell_row - 1,
                                                                                                  cell_col - 1) is not None else 0
            except:
                follow_cell = 0
        # print(re_cell)
        # print(follow_cell)
        if abs(re_cell) < 0.0000000001:
            re_cell = 0
        if abs(follow_cell) < 0.0000000001:
            follow_cell = 0
        if re_cell == 0 and follow_cell == 0:
            out = 2
            check_out.append([reference, '*', '%.2f' % out])
            # continue
        elif re_cell != 0:
            out = follow_cell / re_cell
            check_out.append([reference, '*', '%.2f' % out])
        else:
            check_out.append([reference, '*', 0])
    return check_out


# 用于判断MR关系是否足够构成一个mr
def check_ismr(mr_list, o_list):
    o_set = []
    for o in o_list:
        if o not in o_set:
            o_set.append(o)
    for cell in o_set:
        if o_list.count(cell) / len(o_list) > 0.6:
            for _ in range(len(mr_list)):
                mr_list[_].append(o_list[_])
            break


def check_addmr(ws1, filename2, regions, reference, sheetname, buffer, average_buffer, wb2, in_sheetname):
    ws2 = wb2.sheet_by_name(str(in_sheetname))
    # wb2 = openpyxl.load_workbook(filename2, data_only=True)
    # ws2 = wb2[sheetname]
    check_out = []
    for single in regions:
        cell_row = single[0]
        cell_col = single[1]
        if (cell_row, cell_col) in average_buffer.keys():
            re_cell = average_buffer[(cell_row, cell_col)]
        else:
            try:
                re_cell = float(ws1.cell_value(cell_row - 1, cell_col - 1)) if ws1.cell_value(cell_row - 1,
                                                                                              cell_col - 1) is not None else 0
            except:
                re_cell = 0
        if (cell_row, cell_col) in buffer.keys():
            follow_cell = float(buffer[(cell_row, cell_col)])
        else:
            try:
                follow_cell = float(ws2.cell_value(cell_row - 1, cell_col - 1)) if ws2.cell_value(cell_row - 1,
                                                                                                  cell_col - 1) is not None else 0
            except:
                follow_cell = 0
        # print(cell_row,cell_col)
        # print(re_cell)
        # print(follow_cell)
        out = follow_cell - re_cell
        if abs(out) < 0.00001:
            out = 0.0000
        check_out.append([reference, '+', '%.4f' % out])
    # wb = openpyxl.load_workbook(filename, data_only=False)
    # wb.save(r'E:\pydate\EXCEL\test2(1).xlsx')
    # just_open(r'E:\pydate\EXCEL\test2(1).xlsx')

    return check_out


def domutils_excel(cells_list, other_cells, ws_oriture, sheetname, wb, filename_fuben, sub_region, buffer, ws_orifalse,
                   in_wb, in_sheetname):
    tag = 0
    if str(in_sheetname) not in in_wb.sheetnames:
        in_ws = in_wb.create_sheet(str(in_sheetname))
    else:
        in_ws = in_wb[str(in_sheetname)]
    times = int(len(other_cells) / len(cells_list))
    referenced_cells = []
    for i in range(len(cells_list)):
        if tuple(cells_list[i]) in sub_region:
            tag = 1
            for m in range(len(sub_region)):
                in_ws.cell(row=sub_region[m][0], column=sub_region[m][1]).value = ws_orifalse.cell(row=sub_region[m][0],
                                                                                                   column=sub_region[m][
                                                                                                       1]).value
            in_wb.save(filename_fuben)
            just_open(filename_fuben)
            in_wb1 = openpyxl.load_workbook(filename_fuben, data_only=True)
            in_ws1 = in_wb1[str(in_sheetname)]
            buffer[tuple(cells_list[i])] = in_ws1.cell(row=cells_list[i][0], column=cells_list[i][1]).value
            try:
                ws_oriture_value = ws_oriture.cell(cells_list[i][0] - 1, cells_list[i][1] - 1)
                if ws_oriture_value.ctype is not 0:
                    try:
                        in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = float(
                            ws_oriture_value.value) * 2
                    except:
                        in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = 0
                else:
                    in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = 0
            except:
                in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = 0
        else:
            try:
                ws_oriture_value = ws_oriture.cell(cells_list[i][0] - 1, cells_list[i][1] - 1)
                if ws_oriture_value.ctype is not 0:
                    try:
                        in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = float(
                            ws_oriture_value.value) * 2
                    except:
                        in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = 0
                else:
                    in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = 0
            except:
                in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = 0
        for j in range(times * i, times * (i + 1)):
            if tuple(other_cells[j]) in sub_region:
                tag = 2
                in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = ws_orifalse.cell(
                    row=other_cells[j][0],
                    column=other_cells[j][
                        1]).value
                in_wb.save(filename_fuben)
                just_open(filename_fuben)
                in_wb1 = openpyxl.load_workbook(filename_fuben, data_only=True)
                in_ws1 = in_wb1[str(in_sheetname)]
                buffer[tuple(other_cells[j])] = in_ws1.cell(row=other_cells[j][0], column=other_cells[j][1]).value
                try:
                    ws_oriture_value = ws_oriture.cell(other_cells[j][0] - 1, other_cells[j][1] - 1)
                    if ws_oriture_value.ctype is not 0:
                        in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = ws_oriture_value.value
                    else:
                        in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = 0
                except:
                    in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = 0
            else:
                try:
                    ws_oriture_value = ws_oriture.cell(other_cells[j][0] - 1, other_cells[j][1] - 1)
                    if ws_oriture_value.ctype is not 0:
                        in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = ws_oriture_value.value
                    else:
                        in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = 0
                except:
                    in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = 0
    if tag == 0:
        for m in range(len(sub_region)):
            in_ws.cell(row=sub_region[m][0], column=sub_region[m][1]).value = ws_orifalse.cell(row=sub_region[m][0],
                                                                                               column=sub_region[m][
                                                                                                   1]).value
    elif tag == 2:
        referenced_cells = list(set(sub_region) - set(referenced_cells))
        for m in range(len(referenced_cells)):
            in_ws.cell(row=referenced_cells[m][0], column=referenced_cells[m][1]).value = ws_orifalse.cell(
                row=referenced_cells[m][0],
                column=referenced_cells[m][
                    1]).value


# 用于判断是否为没有乘法 如没有乘法则所有乘一个常数
def check_isnomulti(region, ws_false):
    result = []
    for cell in region:
        cell_value = ws_false.cell(row=cell[0], column=cell[1]).value
        # 判断是否为没有乘法
        if '*' in cell_value:
            result.append(1)
        elif 'PRODUCT' in cell_value:
            result.append(1)
        elif '/' in cell_value:
            result.append(1)
        else:
            result.append(0)
    return result


def doadd_excel(cells_list, other_cells, ws_oriture, sheetname, wb, filename_fuben, sub_region, buffer, ws_orifalse,
                in_wb, in_sheetname):
    if str(in_sheetname) not in in_wb.sheetnames:
        in_ws = in_wb.create_sheet(str(in_sheetname))
    else:
        in_ws = in_wb[str(in_sheetname)]
    tag = 0
    # print(cells_list)
    # print(other_cells)
    add_number = 10
    tag1 = 0
    times = int(len(other_cells) / len(cells_list))
    referenced_cells = []
    for i in range(len(cells_list)):
        if tuple(cells_list[i]) in sub_region:
            tag = 1
            tag1 = 1
            for m in range(len(sub_region)):
                in_ws.cell(row=sub_region[m][0], column=sub_region[m][1]).value = ws_orifalse.cell(row=sub_region[m][0],
                                                                                                   column=sub_region[m][
                                                                                                       1]).value
            in_wb.save(filename_fuben)
            just_open(filename_fuben)
            in_wb1 = openpyxl.load_workbook(filename_fuben, data_only=True)
            in_ws1 = in_wb1[str(in_sheetname)]
            buffer[tuple(cells_list[i])] = in_ws1.cell(row=cells_list[i][0], column=cells_list[i][1]).value
            try:
                ws_oriture_value = ws_oriture.cell(cells_list[i][0] - 1, cells_list[i][1] - 1)
                if ws_oriture_value.ctype is not 0:
                    try:
                        in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = float(
                            ws_oriture_value.value) + add_number
                    except:
                        in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = add_number
                else:
                    in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = add_number
            except:
                in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = add_number

        else:
            try:
                ws_oriture_value = ws_oriture.cell(cells_list[i][0] - 1, cells_list[i][1] - 1)
                if ws_oriture_value.ctype is not 0:
                    try:
                        in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = float(
                            ws_oriture_value.value) + add_number
                    except:
                        in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = add_number
                else:
                    in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = add_number
            except:
                in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value = add_number

        # print(']]]]]]]]')

        # print(in_ws.cell(row=cells_list[i][0], column=cells_list[i][1]).value)
        for j in range(times * i, times * (i + 1)):
            if tuple(other_cells[j]) in sub_region:
                referenced_cells.append(tuple(other_cells[j]))
                tag = 2
                # for m in range(len(sub_region)):
                in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = ws_orifalse.cell(
                    row=other_cells[j][0],
                    column=other_cells[j][
                        1]).value
                in_wb.save(filename_fuben)
                just_open(filename_fuben)
                in_wb1 = openpyxl.load_workbook(filename_fuben, data_only=True)
                in_ws1 = in_wb1[str(in_sheetname)]

                buffer[tuple(other_cells[j])] = in_ws1.cell(row=other_cells[j][0], column=other_cells[j][1]).value
                try:
                    ws_oriture_value = ws_oriture.cell(other_cells[j][0] - 1, other_cells[j][1] - 1)
                    # print('--------------{}'.format(ws_oriture_value.value))
                    if ws_oriture_value.ctype is not 0:
                        in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = float(
                            ws_oriture_value.value)
                    else:
                        in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = None

                except:
                    in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = 0
            else:
                try:
                    ws_oriture_value = ws_oriture.cell(other_cells[j][0] - 1, other_cells[j][1] - 1)
                    # print(ws_oriture_value)
                    if ws_oriture_value.ctype is not 0:
                        in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = ws_oriture_value.value
                    else:
                        in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = None
                except:
                    in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value = 0
    if tag == 0:
        # print('=======================')
        for m in range(len(sub_region)):
            in_ws.cell(row=sub_region[m][0], column=sub_region[m][1]).value = ws_orifalse.cell(row=sub_region[m][0],
                                                                                               column=sub_region[m][
                                                                                                   1]).value
    elif tag == 2:
        # print(sub_region)
        # print(referenced_cells)
        referenced_cells = list(set(sub_region) - set(referenced_cells))
        for m in range(len(referenced_cells)):
            in_ws.cell(row=referenced_cells[m][0], column=referenced_cells[m][1]).value = ws_orifalse.cell(
                row=referenced_cells[m][0],
                column=referenced_cells[m][
                    1]).value
    return tag1
    # print(in_ws.cell(row=other_cells[j][0], column=other_cells[j][1]).value)
    # for other_cell in other_cells:
    #     if ws_oriture.cell(row=other_cell[0], column=other_cell[1]).value is not None:
    #         in_ws.cell(row=other_cell[0], column=other_cell[1]).value = float(
    #             ws_oriture.cell(row=other_cell[0], column=other_cell[1]).value)
    # for ws_cell in cells_list:
    #     if ws_oriture.cell(row=ws_cell[0], column=ws_cell[1]).value is not None:
    #         in_ws.cell(row=ws_cell[0], column=ws_cell[1]).value = float(
    #             ws_oriture.cell(row=ws_cell[0], column=ws_cell[1]).value) + 1


def domutil2_excel(cells_list, other_cells, ws_oriture, sheetname, wb, filename_fuben, sub_region, ws_orifalse, in_wb,
                   in_sheetname):
    # wb.save(filename_fuben)
    # just_open(filename_fuben)
    if str(in_sheetname) not in in_wb.sheetnames:
        in_ws = in_wb.create_sheet(str(in_sheetname))
    else:
        in_ws = in_wb[str(in_sheetname)]
    for other_cell in other_cells:
        ws_oriture_value = ws_oriture.cell(other_cell[0] - 1, other_cell[1] - 1)
        if ws_oriture_value.ctype is not 0:
            try:
                in_ws.cell(row=other_cell[0], column=other_cell[1]).value = ws_oriture_value.value
            except:
                in_ws.cell(row=other_cell[0], column=other_cell[1]).value = 0
        else:
            in_ws.cell(row=other_cell[0], column=other_cell[1]).value = 0
    for ws_cell in cells_list:
        ws_oriture_value = ws_oriture.cell(ws_cell[0] - 1, ws_cell[1] - 1)
        if ws_oriture_value.ctype is not 0:
            try:
                in_ws.cell(row=ws_cell[0], column=ws_cell[1]).value = float(
                    ws_oriture_value.value) * 2
            except:
                in_ws.cell(row=ws_cell[0], column=ws_cell[1]).value = 0
        else:
            in_ws.cell(row=ws_cell[0], column=ws_cell[1]).value = 0
    in_ws.cell(row=sub_region[0][0], column=sub_region[0][1]).value = ws_orifalse.cell(row=sub_region[0][0],
                                                                                       column=sub_region[0][1]).value


def doadd2_excel(cells_list, other_cells, ws_oritrue, sheetname, wb, filename_fuben, sub_region, ws_orifalse, in_wb,
                 in_sheetname):
    # wb.save(filename_fuben)
    # just_open(filename_fuben)

    if str(in_sheetname) not in in_wb.sheetnames:
        in_ws = in_wb.create_sheet(str(in_sheetname))
    else:
        in_ws = in_wb[str(in_sheetname)]
    for other_cell in other_cells:
        ws_oritrue_value = ws_oritrue.cell(other_cell[0] - 1, other_cell[1] - 1)
        if ws_oritrue_value.ctype is not 0:
            in_ws.cell(row=other_cell[0], column=other_cell[1]).value = ws_oritrue_value.value
        else:
            in_ws.cell(row=other_cell[0], column=other_cell[1]).value = 0
    for ws_cell in cells_list:
        ws_oritrue_value = ws_oritrue.cell(ws_cell[0] - 1, ws_cell[1] - 1)
        if ws_oritrue_value.ctype is not 0:
            try:
                in_ws.cell(row=ws_cell[0], column=ws_cell[1]).value = float(
                    ws_oritrue_value.value) + 1
            except:
                in_ws.cell(row=ws_cell[0], column=ws_cell[1]).value = 1
        else:
            in_ws.cell(row=ws_cell[0], column=ws_cell[1]).value = 1
    in_ws.cell(row=sub_region[0][0], column=sub_region[0][1]).value = ws_orifalse.cell(row=sub_region[0][0],
                                                                                       column=sub_region[0][1]).value


def just_open(filename):
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()

    xlBook.Close()


def check_onlymulti(region, ws_false):
    count = 0
    dis_count = 0
    for cell in region:
        if dis_count == 4:
            return 0
        cell_value = ws_false.cell(row=cell[0], column=cell[1]).value
        if 'PRODUCT' in cell_value:
            count += 1
            continue
        result1 = re.match('=\(*[\\\+\-]*\$*[A-Z]+\$*[0-9]+\)*([/\\\*]\(*[\\\+\-]*\$*[A-Z]+\$*[0-9]+\)*)+', cell_value)
        if result1 != None:
            if cell_value == result1.group(0):
                count += 1
        else:
            dis_count += 1
    if count > 0.8 * len(region):
        return 1
    else:
        return 0


def find_mr(filename, filename_fuben, sheetname, example_region, example_reference):
    swap_suspic = []
    suspicion = dict()
    final_region = []
    final_mrs = []
    return_reference = []
    wb = openpyxl.load_workbook(filename, data_only=False)
    ws_false = wb[sheetname]
    wb_true = xlrd.open_workbook(filename)
    ws_true = wb_true.sheet_by_name(sheetname)
    wb.save(filename_fuben)
    # just_open(filename_fuben)
    sum_regions = 0
    un_add_regions = 0
    un_swap_regions = 0
    satisfy_add = 0
    satisfy_swap = 0
    both_satisfy = 0
    perfect_add_region = 0
    perfect_swap_region = 0
    un_apply_add = 0
    for i in range(len(example_region)):

        sub_region = example_region[i]
        sub_reference = example_reference[i]
        set_subreference = []
        for refer in sub_reference:
            if refer not in set_subreference:
                set_subreference.append(refer)
        sub_reference = set_subreference
        for i in range(len(sub_region)):
            for cell in sub_reference:
                if len(cell) == 3:
                    row = cell[0]
                    column = cell[1]
                else:
                    row = sub_region[i][0] + cell[0]
                    column = sub_region[i][1] + cell[1]
                try:
                    # print(row,column)
                    # print(ws_false.cell(row, column).value)
                    if ws_true.cell(row - 1, column - 1).ctype is 0 or ws_true.cell(row - 1,
                                                                                    column - 1).value == 0 and '=' not in str(
                        ws_false.cell(row, column).value):
                        ws_false.cell(row, column).value = random.randint(1, 100)
                except:
                    continue
    index = filename_fuben.rfind('.')
    str_list = list(filename_fuben)
    str_list.insert(index, '(1)')
    excel_path_fuben = ''.join(str_list)
    wb.save(excel_path_fuben)

    try:
        just_open(excel_path_fuben)
    except:
        pass
    wb = openpyxl.load_workbook(excel_path_fuben, data_only=False)
    ws_false = wb[sheetname]
    wb_true = xlrd.open_workbook(excel_path_fuben)
    ws_true = wb_true.sheet_by_name(sheetname)
    for i in range(len(example_region)):
        # print('{}/{}'.format(i, len(example_region)))
        try:

            aa = 0
            in_sheetname = 0
            in_wb = openpyxl.load_workbook(filename_fuben, data_only=False)
            # print(i)
            sub_region = example_region[i]
            sub_reference = example_reference[i]
            set_subreference = []
            for refer in sub_reference:
                if refer not in set_subreference:
                    set_subreference.append(refer)
            sub_reference = set_subreference
            # print(sub_region)
            # print(sub_reference)
            mr_list = [[] for _ in range(len(sub_region))]
            # 对于多参考值单个函数进行mr分析
            big_num = float('-INF')
            max_row = float('-INF')
            min_row = float('INF')
            max_column = float('-INF')
            min_column = float('INF')
            if len(example_region[i]) == 1 and len(example_reference[i]) > 8:
                for reference in example_reference[i]:
                    if reference[0] > max_row:
                        max_row = reference[0]
                    if reference[0] < min_row:
                        min_row = reference[0]
                    if reference[1] > max_column:
                        max_column = reference[1]
                    if reference[1] < min_column:
                        min_column = reference[1]
            # print("-----------")
            # print(max_row, min_row, max_column, min_column)
            if max_row != big_num:
                sum_regions += 1
                # if max_row != min_row and max_column != min_column:
                #     tagname = -1
                if 'IF' in ws_false.cell(row=sub_region[0][0], column=sub_region[0][1]).value:
                    continue
                if abs(max_row - min_row) < abs(max_column - min_column):
                    tagname = 0
                else:
                    tagname = 1
                subregion_list = []
                taglist = []
                for reference in sub_reference:
                    if reference[tagname] not in taglist:
                        taglist.append(reference[tagname])
                        subregion_list.append([])
                        subregion_list[taglist.index(reference[tagname])].append(reference)
                    else:
                        subregion_list[taglist.index(reference[tagname])].append(reference)
                subregion_lists = []
                for _ in subregion_list:
                    # print(_)
                    subregion_lists.extend(_)
                logs_reference = []
                # print('doing forward')
                if str(in_sheetname) not in in_wb.sheetnames:
                    in_ws1 = in_wb.create_sheet(str(in_sheetname))
                else:
                    in_ws1 = in_wb[str(in_sheetname)]
                logs_reference.append([])
                in_sheetname = in_sheetname + 1
                row = sub_region[0][0]
                column = sub_region[0][1]
                for references in subregion_list:
                    value_list = []
                    for reference in references:
                        # print(reference)
                        value = ws_true.cell_value(row + reference[0] - 1, column + reference[1] - 1)
                        value_list.append(value)
                    # 将第一个值放到最后一个
                    first_value = value_list.pop(0)
                    value_list.append(first_value)
                    for j in range(len(references)):
                        in_ws1.cell(row=row + references[j][0], column=column + references[j][1]).value = value_list[j]
                in_ws1.cell(row=row, column=column).value = ws_false.cell(row=row, column=column).value
                forward_sheet = in_sheetname
                # print('doing multi')
                answer_multi = dict()
                reference_list = []
                if len(subregion_list) < 5:
                    for i in range(1, len(subregion_list) + 1):
                        iter_list = list(itertools.combinations(subregion_list, i))
                        for iter in iter_list:
                            reference_list.append(iter)
                else:
                    for i in [1, len(subregion_list)]:
                        iter_list = list(itertools.combinations(subregion_list, i))
                        for iter in iter_list:
                            reference_list.append(iter)
                for references in reference_list:
                    # print(references)
                    cell_list = []
                    other_list = []
                    other_references = [item for item in subregion_list if item not in references]
                    # print(other_references)
                    for reference in references:
                        for sub_reference in reference:
                            # print(sub_reference)
                            if len(sub_reference) == 3:
                                row = sub_reference[0]
                                column = sub_reference[1]
                                cell_list.append([row, column])
                            else:
                                row = sub_reference[0] + sub_region[0][0]
                                column = sub_reference[1] + sub_region[0][1]
                                cell_list.append([row, column])
                    for other_reference in other_references:
                        for sub_other_reference in other_reference:
                            if len(sub_other_reference) == 3:
                                row = sub_other_reference[0]
                                column = sub_other_reference[1]
                                other_list.append([row, column])
                            else:
                                row = sub_other_reference[0] + sub_region[0][0]
                                column = sub_other_reference[1] + sub_region[0][1]
                                other_list.append([row, column])
                    buffer = dict()
                    domutil2_excel(cell_list, other_list, ws_true, sheetname, wb, filename_fuben, sub_region, ws_false,
                                   in_wb, in_sheetname)
                    logs_reference.append(references)
                    in_sheetname = in_sheetname + 1
                mutil_sheetname = in_sheetname
                # 做加法的excel
                # print('doing add')
                answer_add = dict()
                for i in range(len(subregion_list)):
                    cell_list = []
                    other_list = []
                    for cell in subregion_list[i]:
                        # print(cell)
                        # print(sub_region)
                        if len(cell) == 3:
                            row = cell[0]
                            column = cell[1]
                            cell_list.append([row, column])
                        else:
                            row = cell[0] + sub_region[0][0]
                            column = cell[1] + sub_region[0][1]
                            cell_list.append([row, column])
                    for j in range(0, i):
                        for other in subregion_list[j]:
                            if len(other) == 3:
                                row = other[0]
                                column = other[1]
                                other_list.append([row, column])
                            else:
                                row = other[0] + sub_region[0][0]
                                column = other[1] + sub_region[0][1]
                                other_list.append([row, column])
                    for j in range(i + 1, len(subregion_list)):
                        for other in subregion_list[j]:
                            if len(other) == 3:
                                row = other[0]
                                column = other[1]
                                other_list.append([row, column])
                            else:
                                row = other[0] + sub_region[0][0]
                                column = other[1] + sub_region[0][1]
                                other_list.append([row, column])
                    buffer = dict()
                    doadd2_excel(cell_list, other_list, ws_true, sheetname, wb, filename_fuben, sub_region, ws_false,
                                 in_wb, in_sheetname)
                    logs_reference.append(i)
                    in_sheetname = in_sheetname + 1
                add_sheetname = in_sheetname

                # 对in_wb 保存更改
                in_wb.save(filename_fuben)
                try:
                    just_open(filename_fuben)
                except:
                    continue
                # check 结果
                in_wb = openpyxl.load_workbook(filename_fuben, data_only=True)
                for sheetname_num in range(0, forward_sheet):
                    in_ws = in_wb[str(sheetname_num)]
                    row = sub_region[0][0]
                    column = sub_region[0][1]
                    origin_value = ws_true.cell_value(row - 1, column - 1)
                    changed_value = in_ws.cell(row=row, column=column).value
                    # print('======')
                    # print(origin_value,changed_value)
                    if origin_value == changed_value:
                        is_forward = True
                    else:
                        is_forward = False
                for sheetname_num in range(forward_sheet, mutil_sheetname):
                    ws_fuben = in_wb[str(sheetname_num)]
                    references = logs_reference[sheetname_num]
                    a = float(ws_fuben.cell(row=sub_region[0][0], column=sub_region[0][1]).value) / (
                        float(ws_true.cell_value(
                            sub_region[0][0] - 1,
                            sub_region[0][
                                1] - 1)) if float(
                            ws_true.cell_value(sub_region[0][0] - 1, sub_region[0][1] - 1)) != 0 else 1)
                    # print(a)
                    if a % 2 == 0 or a % 0.5 == 0:
                        if a in answer_multi.keys():
                            answer_multi[a].append(references)
                        else:
                            answer_multi[a] = []
                            answer_multi[a].append(references)
                for sheetname_num in range(mutil_sheetname, add_sheetname):
                    ws_fuben = in_wb[str(sheetname_num)]
                    i = logs_reference[sheetname_num]
                    b = ws_fuben.cell(row=sub_region[0][0], column=sub_region[0][1]).value - ws_true.cell_value(
                        sub_region[0][0] - 1,
                        sub_region[0][
                            1] - 1)
                    if len(subregion_list[i][0]) == 3:
                        first_cell_row = subregion_list[i][0][0]
                        first_cell_column = get_column_letter(subregion_list[i][0][1])
                    else:
                        first_cell_row = subregion_list[i][0][0] + sub_region[0][0]
                        first_cell_column = get_column_letter(subregion_list[i][0][1] + sub_region[0][1])
                    if len(subregion_list[i][-1]) == 3:
                        last_cell_row = subregion_list[i][-1][0]
                        last_cell_column = get_column_letter(subregion_list[i][-1][1])
                    else:
                        last_cell_row = subregion_list[i][-1][0] + sub_region[0][0]
                        last_cell_column = get_column_letter(subregion_list[i][-1][1] + sub_region[0][1])
                    range_cell = '{}{}:{}{}'.format(first_cell_column, first_cell_row, last_cell_column, last_cell_row)
                    answer_add[range_cell] = b
                # print(answer_add)
                final_region.append(sub_region)
                final_mrs.append([is_forward, answer_multi, answer_add])
                return_reference.append(subregion_lists)
                # 相同value 换位置
                continue
            cell1_row = sub_region[0][0]
            cell1_column = sub_region[0][1]
            if 'MAX' in ws_false.cell(row=cell1_row, column=cell1_column).value or 'MIN' in ws_false.cell(row=cell1_row,
                                                                                                          column=cell1_column).value or 'VLOOKUP' in ws_false.cell(
                row=cell1_row,
                column=cell1_column).value or 'LOOKUP' in ws_false.cell(
                row=cell1_row, column=cell1_column).value or 'OFFSET' in ws_false.cell(row=cell1_row,
                                                                                       column=cell1_column).value or 'IF' in ws_false.cell(
                row=cell1_row, column=cell1_column).value:
                final_region.append(example_region[i])
                final_mrs.append([])
                return_reference.append(example_reference[i])
                sum_regions +=1
                continue
            average_buffer = dict()
            if 'AVERAGE' in ws_false.cell(row=cell1_row, column=cell1_column).value:
                # 进行预处理
                for subsub_reference in sub_reference:
                    for single_cell in sub_region:
                        if len(subsub_reference) == 3:
                            row = subsub_reference[0]
                            column = subsub_reference[1]
                        else:
                            row = subsub_reference[0] + single_cell[0]
                            column = subsub_reference[1] + single_cell[1]
                        if ws_false.cell(row=row, column=column).value is None:
                            ws_false.cell(row=row, column=column).value = 0
                wb.save(filename_fuben)
                just_open(filename_fuben)
                wb_average = openpyxl.load_workbook(filename_fuben, data_only=True)
                ws_average = wb_average[sheetname]
                for single_cell in sub_region:
                    try:

                        average_buffer[(single_cell[0], single_cell[1])] = float(ws_average.cell(row=single_cell[0],
                                                                                                 column=single_cell[
                                                                                                     1]).value)
                    except:
                        average_buffer[(single_cell[0], single_cell[1])] = 0
            logs_reference = []

            # print('do little add')
            sum_buffer = []
            is_only_multi = check_onlymulti(sub_region, ws_false)
            if is_only_multi == 0:
                for i in range(len(sub_reference)):
                    reference = sub_reference[i]
                    other_reference = []  # 除了当前reference之外的列表
                    for _ in range(0, i):
                        other_reference.append(sub_reference[_])
                    for _ in range(i + 1, len(sub_reference)):
                        other_reference.append(sub_reference[_])
                    cell_list = []  # 参考的cells
                    other_celllists = []  # 其他cells
                    for single_cell in sub_region:
                        if len(reference) == 3:
                            row = reference[0]
                            col = reference[1]
                            cell_list.append([row, col])
                        else:
                            row = single_cell[0] + reference[0]
                            col = single_cell[1] + reference[1]
                            cell_list.append([row, col])
                        for other_re in other_reference:
                            if len(other_re) == 3:
                                row = other_re[0]
                                col = other_re[1]
                                other_celllists.append([row, col])
                            else:
                                oth_row = single_cell[0] + other_re[0]
                                oth_col = single_cell[1] + other_re[1]
                                other_celllists.append([oth_row, oth_col])
                    # print(cell_list)
                    # print(other_celllists)
                    buffer = dict()
                    tag1 = doadd_excel(cell_list, other_celllists, ws_true, sheetname, wb, filename_fuben, sub_region,
                                       buffer,
                                       ws_false, in_wb, in_sheetname)
                    if tag1 == 1:
                        aa = 1
                    sum_buffer.append(buffer)
                    # print(buffer)
                    logs_reference.append(reference)
                    in_sheetname = in_sheetname + 1
            add_sheetname = in_sheetname
            # swap
            # print('doing swap')
            # forward==1 代表向下翻转一格 forward==0代表向上翻转一格
            forward = 1
            swap_cell_list = []
            for i in range(len(sub_region)):
                for swap_sub_reference in sub_reference:
                    if len(swap_sub_reference) != 3:
                        swap_cell_row = sub_region[i][0] + swap_sub_reference[0]
                        swap_cell_column = sub_region[i][1] + swap_sub_reference[1]
                        swap_cell_list.append((swap_cell_row, swap_cell_column))
            if aa == 0:
                is_swap = swap_excel(sub_region, sub_reference, ws_true, sheetname, wb, filename_fuben, ws_false, in_wb,
                                     in_sheetname,
                                     forward, swap_cell_list)
                in_sheetname = in_sheetname + 1
            forward = 0
            if aa == 0:
                is_swap = swap_excel(sub_region, sub_reference, ws_true, sheetname, wb, filename_fuben, ws_false, in_wb,
                                     in_sheetname,
                                     forward, swap_cell_list)

            # 保存并读取
            in_wb.save(filename_fuben)
            try:
                just_open(filename_fuben)
            except:
                continue
            in_wb = xlrd.open_workbook(filename_fuben)
            # print('=================', add_sheetname)
            in_un_swap_region = 0
            for sheetname_num in range(0, add_sheetname):
                reference = logs_reference[sheetname_num]
                buffer = sum_buffer[sheetname_num]
                # print(buffer)
                out_list = check_addmr(ws_true, filename_fuben, sub_region, reference, sheetname, buffer,
                                       average_buffer, in_wb, sheetname_num)
                check_ismr(mr_list, out_list)
            if aa == 0 and is_swap == 1:
                forward = 1
                sheetname_num = add_sheetname
                different_index_forward = check_swap(ws_true, filename, sub_region, sheetname, in_wb, sheetname_num,
                                                     forward)
                forward = forward - 1
                sheetname_num += 1
                different_index_backward = check_swap(ws_true, filename, sub_region, sheetname, in_wb, sheetname_num,
                                                      forward)
                if list(set(different_index_forward) & set(different_index_backward)):
                    in_un_swap_region = 1
                swap_suspic.append(list(set(different_index_forward) & set(different_index_backward)))
            mrs = []
            splite_region = []
            # for _ in mr_list:
            #     print(_)
            for index in range(len(mr_list)):
                if mr_list[index] not in mrs:
                    mrs.append(mr_list[index])
                    splite_region.append([])
                    splite_region[len(splite_region) - 1].append(sub_region[index])
                else:
                    in_index = mrs.index(mr_list[index])
                    splite_region[in_index].append(sub_region[index])
            # print(mrs)
            # print(splite_region)
            # print(len(mrs))
            # print('--------------------------------------------------')
            in_un_add_region = 0
            in_perferenct_add = 0
            in_un_apply_add = 0
            if len(mrs) < 8:
                in_sum_region = 1
                if mrs[0] ==[]:
                    in_un_apply_add = 1
                if len(mrs) == 1 :
                    in_perferenct_add = 1
                for mr in mrs:
                    final_mrs.append(mr)
                    return_reference.append(sub_reference)
                for _ in splite_region:
                    final_region.append(_)
                if find_suspicion(splite_region, mrs):
                    suspicion_list, suspicion_num = find_suspicion(splite_region, mrs)
                    # print(suspicion_list)
                    if suspicion_list:
                        in_un_add_region = 1
                        for sus_index, sub_sus in enumerate(suspicion_list):
                            for sub_sus_index in sub_sus:
                                suspicion[sub_sus_index] = suspicion_num[sus_index]
            else:
                # print('-0-0-0-0-0-0-0')
                in_sum_region = 2
                max_len = 0
                index = -1
                for i, sub_splite_region in enumerate(splite_region):
                    if max_len < len(sub_splite_region):
                        max_len = len(sub_splite_region)
                        index = i
                max_splite = splite_region.pop(index)
                max_mr = mrs.pop(index)
                final_mrs.append(max_mr)
                final_region.append(max_splite)
                return_reference.append(sub_reference)
                loop_len = len(mrs[1])
                i_tag = 0
                for i in range(0, loop_len):
                    mr_set = []
                    mr_list = []
                    tag = 0

                    for mr in mrs:
                        mr_list.append(mr[i_tag])
                        if mr[i_tag] not in mr_set:
                            mr_set.append(mr[i_tag])
                    for single_mr in mr_set:
                        # print(mr_list.count(single_mr))
                        # print(len(mr_list))
                        if mr_list.count(single_mr) / len(mr_list) > 0.6:
                            tag = 1
                            i_tag = i_tag + 1
                            break
                    if tag == 0:
                        for mr in mrs:
                            mr.pop(i_tag)
                mrs_sec = []
                splite_region_sec = []
                # for _ in mr_list:
                #     print(_)
                for index in range(len(mrs)):
                    if mrs[index] not in mrs_sec:
                        mrs_sec.append(mrs[index])
                        splite_region_sec.append([])
                        for _ in splite_region[index]:
                            splite_region_sec[len(splite_region_sec) - 1].append(_)
                    else:
                        in_index = mrs_sec.index(mrs[index])
                        for _ in splite_region[index]:
                            splite_region_sec[in_index].append(_)

                for mr in mrs_sec:
                    final_mrs.append(mr)
                    return_reference.append(sub_reference)
                for _ in splite_region_sec:
                    final_region.append(_)
                if find_suspicion(splite_region_sec, mrs_sec):
                    suspicion_list, suspicion_num = find_suspicion(splite_region_sec, mrs_sec)
                    if suspicion_list:
                        in_un_add_region = 1
                        for sus_index, sub_sus in enumerate(suspicion_list):
                            for sub_sus_index in sub_sus:
                                suspicion[sub_sus_index] = suspicion_num[sus_index]
            sum_regions = sum_regions + in_sum_region
            if in_un_apply_add != 1:
                un_add_regions = un_add_regions + in_un_add_region
            un_swap_regions += in_un_swap_region
            un_apply_add += in_un_apply_add
            if in_un_add_region == 0 and in_un_apply_add != 1:
                satisfy_add += 1
            if in_un_swap_region == 0:
                satisfy_swap += 1
            if in_un_swap_region == 0 and in_un_add_region == 0 and mrs[0] != []:
                both_satisfy += 1
            if in_un_apply_add!=1:
                perfect_add_region += in_perferenct_add

        # print('secend')
        # print(mrs_sec)
        # print(splite_region_sec)
        # print(len(mrs_sec))

        except:
            final_region.append(example_region[i])
            final_mrs.append([])
            return_reference.append(example_reference[i])
    # print(final_mrs)
    # print(final_region)
    # sus_mr_dict = mr_suspicion(final_region, final_mrs)
    # print(suspicion)
    # 将两个可疑值字典合并
    # for key, value in sus_mr_dict.items():
    #     if key in suspicion.keys():
    #         suspicion[key] = suspicion[key] + value
    #     else:
    #         suspicion[key] = value
    # for key in sorted(sus_mr_dict, key=sus_mr_dict.__getitem__, reverse=True):
    #     if key not in suspicion.keys():
    #         suspicion[key] = sus_mr_dict[key]
    # print(sus_mr_dict)

    # print(len(final_region))
    # for mr in final_mrs:
    #     print(mr)
    final_check(final_region, suspicion, ws_true)
    os.remove(excel_path_fuben)
    final_swap = []
    for _ in swap_suspic:
        for sus_cell in _:
            final_swap.append(sus_cell)
            if sus_cell not in suspicion.keys():
                suspicion[sus_cell] = 0.02
    # print(suspicion)
    # print(sum_regions, un_add_regions, un_swap_regions, perfect_add_region, both_satisfy)
    if (un_add_regions+satisfy_add) != (un_swap_regions+satisfy_swap):
        print(filename)
    number_list = [un_apply_add, un_add_regions, un_swap_regions, satisfy_add, satisfy_swap, both_satisfy,
                   perfect_add_region]
    return final_region, final_mrs, return_reference, suspicion, final_swap, number_list

    # color_cells.color_cell(final_region, filename_fuben, sheetname)


xlApp.Quit()
